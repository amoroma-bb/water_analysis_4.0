from tkinter import*
from tkinter import ttk
import tkinter.font as tkFont
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import sqlite3
from typing import Text

def main():
    root = Tk()
    root.title('Sendo Water Analysis 4.0')
    
    # Font
    ft = tkFont.Font(family='Consolas', size=10, weight=tkFont.NORMAL)
    ft1 = tkFont.Font(family='Consolas', size=10, weight=tkFont.NORMAL)
    ft2 = tkFont.Font(family='Consolas', size=10, weight=tkFont.NORMAL)
    # Frame
    
    wrapper = LabelFrame(root,text ='test wrapper 1')
    wrapper2 = LabelFrame(root,text='test wrapper 2')
    wrapper.pack(padx=100,pady=50,fill = 'x',expand='no')
    wrapper2.pack(padx=100,pady=50,fill = 'x',expand='no')
    #wrapper = ttk.Frame(root,padding='3 3 12 12')
    #wrapper.grid(column=0,row=0,sticky=(N, W, E, S))
    #wrapper.columnconfigure(0,weight=1)
    #wrapper.rowconfigure(0,weight=1)
    #
    label_xm = ttk.Label(wrapper,text='Items',font=ft)
    label_xm.grid(column=0,row=2,sticky=(W))
    # Label Ca
    label_ca = ttk.Label(wrapper, text='Ca',font=ft)
    label_ca.grid(column=0,row=3,sticky=(W))
    # Label hardness
    label_yd = ttk.Label(wrapper, text='Hardness',font=ft)
    label_yd.grid(column=0,row=4,sticky=(W))
    # Label alkalinity
    label_jd = ttk.Label(wrapper,text='Alkalinity',font=ft)
    label_jd.grid(column=0,row=5,sticky=(W))
    # Label Cl-
    label_cl = ttk.Label(wrapper, text='Cl-',font=ft)
    label_cl.grid(column=0,row=6,sticky=(W))
    # Label total phosphorus
    label_p = ttk.Label(wrapper,text='Total P',font=ft)
    label_p.grid(column=0,row=7,sticky=(W))

    # Laber for water items
    label_water1 =ttk.Label(wrapper,text=' No. 1 ',font=ft)
    label_water1.grid(column=1, row=0, sticky=(N))
    label_water2 =ttk.Label(wrapper,text=' No. 2 ',font=ft)
    label_water2.grid(column=3, row=0, sticky=(N))
    label_water3 =ttk.Label(wrapper,text=' No. 3 ',font=ft)
    label_water3.grid(column=5, row=0, sticky=(N))
    label_water4 =ttk.Label(wrapper,text=' No. 4 ',font=ft)
    label_water4.grid(column=7, row=0, sticky=(N))
    label_water5 =ttk.Label(wrapper,text=' No. 5 ',font=ft)
    label_water5.grid(column=9, row=0, sticky=(N))
    label_water6 =ttk.Label(wrapper,text=' No. 6 ',font=ft)
    label_water6.grid(column=11,row=0, sticky=(N))
    

    # Default value for water entry
    v1 = StringVar(root, value='0')
    v2 = StringVar(root, value='0')
    v3 = StringVar(root, value='0')
    v4 = StringVar(root, value='0')
    v5 = StringVar(root, value='0')
    v6 = StringVar(root, value='0')
    v7 = StringVar(root, value='0')
    v8 = StringVar(root, value='0')
    v11 = StringVar(root, value='0')
    v12 = StringVar(root, value='0')
    v13 = StringVar(root, value='0')
    v14 = StringVar(root, value='0')
    v15 = StringVar(root, value='0')
    v16 = StringVar(root, value='0')
    v17 = StringVar(root, value='0')
    v18 = StringVar(root, value='0')
    v21 = StringVar(root, value='0')
    v22 = StringVar(root, value='0')
    v23 = StringVar(root, value='0')
    v24 = StringVar(root, value='0')
    v25 = StringVar(root, value='0')
    v26 = StringVar(root, value='0')
    v27 = StringVar(root, value='0')
    v28 = StringVar(root, value='0')
    v31 = StringVar(root, value='0')
    v32 = StringVar(root, value='0')
    v33 = StringVar(root, value='0')
    v34 = StringVar(root, value='0')
    v35 = StringVar(root, value='0')
    v36 = StringVar(root, value='0')
    # Default value for water volume
    vw_ca1 = StringVar(root, value='15')
    vw_ca2 = StringVar(root, value='15')
    vw_ca3 = StringVar(root, value='15')
    vw_ca4 = StringVar(root, value='15')
    vw_ca5 = StringVar(root, value='15')
    vw_ca6 = StringVar(root, value='15')
    vw_hn1 = StringVar(root, value='15')
    vw_hn2 = StringVar(root, value='15')
    vw_hn3 = StringVar(root, value='15')
    vw_hn4 = StringVar(root, value='15')
    vw_hn5 = StringVar(root, value='15')
    vw_hn6 = StringVar(root, value='15')
    vw_alk1 = StringVar(root, value='50')
    vw_alk2 = StringVar(root, value='50')
    vw_alk3 = StringVar(root, value='50')
    vw_alk4 = StringVar(root, value='50')
    vw_alk5 = StringVar(root, value='50')
    vw_alk6 = StringVar(root, value='50')
    vw_cl1 = StringVar(root, value='15')
    vw_cl2 = StringVar(root, value='15')
    vw_cl3 = StringVar(root, value='15')
    vw_cl4 = StringVar(root, value='15')
    vw_cl5 = StringVar(root, value='15')
    vw_cl6 = StringVar(root, value='15')
    vw_tp1 = StringVar(root, value='15')
    vw_tp2 = StringVar(root, value='15')
    vw_tp3 = StringVar(root, value='15')
    vw_tp4 = StringVar(root, value='15')
    vw_tp5 = StringVar(root, value='15')
    vw_tp6 = StringVar(root, value='15')

    excel_path = StringVar(root,value='')

    # Entry for water items
    entry_water1 = ttk.Entry(wrapper, width=5)
    entry_water1.grid(column=1,row=2,sticky=(N))
    entry_water2 = ttk.Entry(wrapper, width=5,textvariable=v31)
    entry_water2.grid(column=3,row=2,sticky=(N))
    entry_water3 = ttk.Entry(wrapper, width=5,textvariable=v1)
    entry_water3.grid(column=5,row=2,sticky=(N))
    entry_water4 = ttk.Entry(wrapper, width=5,textvariable=v2)
    entry_water4.grid(column=7,row=2,sticky=(N))
    entry_water5 = ttk.Entry(wrapper, width=5,textvariable=v3)
    entry_water5.grid(column=9,row=2,sticky=(N))
    entry_water6 = ttk.Entry(wrapper, width=5,textvariable=v4)
    entry_water6.grid(column=11,row=2,sticky=(N))

    # Entry for Ca
    entry_ca1 = ttk.Entry(wrapper, width=5)
    entry_ca1.grid(column=1,row=3,sticky=(N))
    entry_ca2 = ttk.Entry(wrapper, width=5,textvariable=v32)
    entry_ca2.grid(column=3,row=3,sticky=(N))
    entry_ca3 = ttk.Entry(wrapper, width=5,textvariable=v5)
    entry_ca3.grid(column=5,row=3,sticky=(N))
    entry_ca4 = ttk.Entry(wrapper, width=5,textvariable=v6)
    entry_ca4.grid(column=7,row=3,sticky=(N))
    entry_ca5 = ttk.Entry(wrapper, width=5,textvariable=v7)
    entry_ca5.grid(column=9,row=3,sticky=(N))
    entry_ca6 = ttk.Entry(wrapper, width=5,textvariable=v8)
    entry_ca6.grid(column=11,row=3,sticky=(N))

    # Entry for hardness
    entry_hn1 = ttk.Entry(wrapper, width=5)
    entry_hn1.grid(column=1,row=4,sticky=(N))
    entry_hn2 = ttk.Entry(wrapper, width=5,textvariable=v33)
    entry_hn2.grid(column=3,row=4,sticky=(N))
    entry_hn3 = ttk.Entry(wrapper, width=5,textvariable=v11)
    entry_hn3.grid(column=5,row=4,sticky=(N))
    entry_hn4 = ttk.Entry(wrapper, width=5,textvariable=v12)
    entry_hn4.grid(column=7,row=4,sticky=(N))
    entry_hn5 = ttk.Entry(wrapper, width=5,textvariable=v13)
    entry_hn5.grid(column=9,row=4,sticky=(N))
    entry_hn6 = ttk.Entry(wrapper, width=5,textvariable=v14)
    entry_hn6.grid(column=11,row=4,sticky=(N))

    # Entry for alkalinity
    entry_alk1 = ttk.Entry(wrapper, width=5)
    entry_alk1.grid(column=1,row=5,sticky=(N))
    entry_alk2 = ttk.Entry(wrapper, width=5,textvariable=v34)
    entry_alk2.grid(column=3,row=5,sticky=(N))
    entry_alk3 = ttk.Entry(wrapper, width=5,textvariable=v15)
    entry_alk3.grid(column=5,row=5,sticky=(N))
    entry_alk4 = ttk.Entry(wrapper, width=5,textvariable=v16)
    entry_alk4.grid(column=7,row=5,sticky=(N))
    entry_alk5 = ttk.Entry(wrapper, width=5,textvariable=v17)
    entry_alk5.grid(column=9,row=5,sticky=(N))
    entry_alk6 = ttk.Entry(wrapper, width=5,textvariable=v18)
    entry_alk6.grid(column=11,row=5,sticky=(N))

    # Entry for cl-
    entry_cl1 = ttk.Entry(wrapper, width=5)
    entry_cl1.grid(column=1,row=6,sticky=(N))
    entry_cl2 = ttk.Entry(wrapper, width=5,textvariable=v35)
    entry_cl2.grid(column=3,row=6,sticky=(N))
    entry_cl3 = ttk.Entry(wrapper, width=5,textvariable=v21)
    entry_cl3.grid(column=5,row=6,sticky=(N))
    entry_cl4 = ttk.Entry(wrapper, width=5,textvariable=v22)
    entry_cl4.grid(column=7,row=6,sticky=(N))
    entry_cl5 = ttk.Entry(wrapper, width=5,textvariable=v23)
    entry_cl5.grid(column=9,row=6,sticky=(N))
    entry_cl6 = ttk.Entry(wrapper, width=5,textvariable=v24)
    entry_cl6.grid(column=11,row=6,sticky=(N))

    # Entry for total phosphorus
    entry_tp1 = ttk.Entry(wrapper, width=5)
    entry_tp1.grid(column=1,row=7,sticky=(N))
    entry_tp2 = ttk.Entry(wrapper, width=5,textvariable=v36)
    entry_tp2.grid(column=3,row=7,sticky=(N))
    entry_tp3 = ttk.Entry(wrapper, width=5,textvariable=v25)
    entry_tp3.grid(column=5,row=7,sticky=(N))
    entry_tp4 = ttk.Entry(wrapper, width=5,textvariable=v26)
    entry_tp4.grid(column=7,row=7,sticky=(N))
    entry_tp5 = ttk.Entry(wrapper, width=5,textvariable=v27)
    entry_tp5.grid(column=9,row=7,sticky=(N))
    entry_tp6 = ttk.Entry(wrapper, width=5,textvariable=v28)
    entry_tp6.grid(column=11,row=7,sticky=(N))
    
    # Entry for volume
    entry_vm_ca1 = ttk.Entry(wrapper,width=4,textvariable=vw_ca1)
    entry_vm_ca1.grid(column=2,row=3,sticky=(N))
    entry_vm_ca2 = ttk.Entry(wrapper,width=4,textvariable=vw_ca2)
    entry_vm_ca2.grid(column=4,row=3,sticky=(N))
    entry_vm_ca3 = ttk.Entry(wrapper,width=4,textvariable=vw_ca3)
    entry_vm_ca3.grid(column=6,row=3,sticky=(N))
    entry_vm_ca4 = ttk.Entry(wrapper,width=4,textvariable=vw_ca4)
    entry_vm_ca4.grid(column=8,row=3,sticky=(N))
    entry_vm_ca5 = ttk.Entry(wrapper,width=4,textvariable=vw_ca5)
    entry_vm_ca5.grid(column=10,row=3,sticky=(N))
    entry_vm_ca6 = ttk.Entry(wrapper,width=4,textvariable=vw_ca6)
    entry_vm_ca6.grid(column=12,row=3,sticky=(N))
    #
    entry_vm_hn1 = ttk.Entry(wrapper,width=4,textvariable=vw_hn1)
    entry_vm_hn1.grid(column=2,row=4,sticky=(N))
    entry_vm_hn2 = ttk.Entry(wrapper,width=4,textvariable=vw_hn2)
    entry_vm_hn2.grid(column=4,row=4,sticky=(N))
    entry_vm_hn3 = ttk.Entry(wrapper,width=4,textvariable=vw_hn3)
    entry_vm_hn3.grid(column=6,row=4,sticky=(N))
    entry_vm_hn4 = ttk.Entry(wrapper,width=4,textvariable=vw_hn4)
    entry_vm_hn4.grid(column=8,row=4,sticky=(N))
    entry_vm_hn5 = ttk.Entry(wrapper,width=4,textvariable=vw_hn5)
    entry_vm_hn5.grid(column=10,row=4,sticky=(N))
    entry_vm_hn6 = ttk.Entry(wrapper,width=4,textvariable=vw_hn6)
    entry_vm_hn6.grid(column=12,row=4,sticky=(N))
    #
    entry_vm_alk1 = ttk.Entry(wrapper,width=4,textvariable=vw_alk1)
    entry_vm_alk1.grid(column=2,row=5,sticky=(N))
    entry_vm_alk2 = ttk.Entry(wrapper,width=4,textvariable=vw_alk2)
    entry_vm_alk2.grid(column=4,row=5,sticky=(N))
    entry_vm_alk3 = ttk.Entry(wrapper,width=4,textvariable=vw_alk3)
    entry_vm_alk3.grid(column=6,row=5,sticky=(N))
    entry_vm_alk4 = ttk.Entry(wrapper,width=4,textvariable=vw_alk4)
    entry_vm_alk4.grid(column=8,row=5,sticky=(N))
    entry_vm_alk5 = ttk.Entry(wrapper,width=4,textvariable=vw_alk5)
    entry_vm_alk5.grid(column=10,row=5,sticky=(N))
    entry_vm_alk6 = ttk.Entry(wrapper,width=4,textvariable=vw_alk6)
    entry_vm_alk6.grid(column=12,row=5,sticky=(N))
    #
    entry_vm_cl1 = ttk.Entry(wrapper,width=4,textvariable=vw_cl1)
    entry_vm_cl1.grid(column=2,row=6,sticky=(N))
    entry_vm_cl2 = ttk.Entry(wrapper,width=4,textvariable=vw_cl2)
    entry_vm_cl2.grid(column=4,row=6,sticky=(N))
    entry_vm_cl3 = ttk.Entry(wrapper,width=4,textvariable=vw_cl3)
    entry_vm_cl3.grid(column=6,row=6,sticky=(N))
    entry_vm_cl4 = ttk.Entry(wrapper,width=4,textvariable=vw_cl4)
    entry_vm_cl4.grid(column=8,row=6,sticky=(N))
    entry_vm_cl5 = ttk.Entry(wrapper,width=4,textvariable=vw_cl5)
    entry_vm_cl5.grid(column=10,row=6,sticky=(N))
    entry_vm_cl6 = ttk.Entry(wrapper,width=4,textvariable=vw_cl6)
    entry_vm_cl6.grid(column=12,row=6,sticky=(N))
    #
    entry_vm_tp1 = ttk.Entry(wrapper,width=4,textvariable=vw_tp1)
    entry_vm_tp1.grid(column=2,row=7,sticky=(N))
    entry_vm_tp2 = ttk.Entry(wrapper,width=4,textvariable=vw_tp2)
    entry_vm_tp2.grid(column=4,row=7,sticky=(N))
    entry_vm_tp3 = ttk.Entry(wrapper,width=4,textvariable=vw_tp3)
    entry_vm_tp3.grid(column=6,row=7,sticky=(N))
    entry_vm_tp4 = ttk.Entry(wrapper,width=4,textvariable=vw_tp4)
    entry_vm_tp4.grid(column=8,row=7,sticky=(N))
    entry_vm_tp5 = ttk.Entry(wrapper,width=4,textvariable=vw_tp5)
    entry_vm_tp5.grid(column=10,row=7,sticky=(N))
    entry_vm_tp6 = ttk.Entry(wrapper,width=4,textvariable=vw_tp6)
    entry_vm_tp6.grid(column=12,row=7,sticky=(N))

    # Entry for path
    entry_path = ttk.Entry(wrapper2,width=20,textvariable=excel_path)
    entry_path.grid(column=1,row=0,ipadx=14,ipady=1) 
    
    # Default con
    with open('standard_concentration.txt','r') as f:
        for line in f.readlines():
            l = line.split(',')
            ca_hn_value = float(l[0])
            alk_value = float(l[1])
            cl_value = float(l[2])
            tp_value = float(l[3])
    # Entry for concentration
    entry_ca_con = ttk.Entry(wrapper2,width=20)
    entry_ca_con.grid(column=1,row=1,ipadx=14,ipady=1)
    entry_alk_con = ttk.Entry(wrapper2,width=20)
    entry_alk_con.grid(column=1,row=2,ipadx=14,ipady=1)
    entry_cl_con = ttk.Entry(wrapper2,width=20)
    entry_cl_con.grid(column=1,row=3,ipadx=14,ipady=1)
    entry_tp_con = ttk.Entry(wrapper2,width=20)
    entry_tp_con.grid(column=1,row=4,ipadx=14,ipady=1)
    
    # Label for concentration
    label_ca_con = ttk.Label(wrapper2,text='Modify EDTA to：',font=ft1)
    label_ca_con.grid(column=0,row=1,sticky=(W))
    label_alk_con = ttk.Label(wrapper2,text='Modify HCl to：',font=ft1)
    label_alk_con.grid(column=0,row=2,sticky=(W))
    label_cl_con = ttk.Label(wrapper2,text='Modify AgNO3 to：',font=ft1)
    label_cl_con.grid(column=0,row=3,sticky=(W))
    label_tp_con = ttk.Label(wrapper2,text='Modify Absorbance to：',font=ft1)
    label_tp_con.grid(column=0,row=4,sticky=(W))

    # Button for change concentration
    button_change = ttk.Button(wrapper2,text='Modify',width=20)
    button_change.grid(column=1,row=5,ipadx=14,ipady=1)
    button_change['command'] = lambda:change(entry_ca_con,entry_alk_con,entry_cl_con,entry_tp_con)

   

    # Label for path
    label_path = ttk.Label(wrapper2,text='Enter the project name',font=ft1)
    label_path.grid(column=0,row=0,sticky=(W),ipadx=12,ipady=1)
    # Label for Tips
    label_tip1 = ttk.Label(wrapper2,text='1.---------',font=ft2)
    label_tip1.grid(column=5,row=1,columnspan=8,sticky=(W))
    label_tip2 = ttk.Label(wrapper2,text='2.---------',font=ft2)
    label_tip2.grid(column=5,row=2,columnspan=8,sticky=(W))
    label_tip3 = ttk.Label(wrapper2,text='3.---------',font=ft2)
    label_tip3.grid(column=5,row=3,columnspan=8,sticky=(W))
    label_tip4 = ttk.Label(wrapper2,text='4.---------',font=ft2)
    label_tip4.grid(column=5,row=4,columnspan=8,sticky=(W))
    label_tip5 = ttk.Label(wrapper2,text='5.---------',font=ft2)
    label_tip5.grid(column=5,row=5,columnspan=8,sticky=(W))
    
    # Button for calculate
    label_space = ttk.Label(wrapper2,text='         ',width=10)
    label_space.grid(column=4,row=0)
    button_cal = ttk.Button(wrapper2,text='Calculate',width=20)
    button_cal.grid(column=5,row=0,sticky=(W))
    button_cal['command'] = lambda:cclt(entry_water1,entry_water2,entry_water3,entry_water4,entry_water5,entry_water6,
    entry_ca1,entry_vm_ca1,entry_ca2,entry_vm_ca2,entry_ca3,entry_vm_ca3,
    entry_ca4,entry_vm_ca4,entry_ca5,entry_vm_ca5,entry_ca6,entry_vm_ca6,
    entry_hn1,entry_vm_hn1,entry_hn2,entry_vm_hn2,entry_hn3,entry_vm_hn3,
    entry_hn4,entry_vm_hn4,entry_hn5,entry_vm_hn5,entry_hn6,entry_vm_hn6,
    entry_alk1,entry_vm_alk1,entry_alk2,entry_vm_alk2,entry_alk3,entry_vm_alk3,
    entry_alk4,entry_vm_alk4,entry_alk5,entry_vm_alk5,entry_alk6,entry_vm_alk6,
    entry_cl1,entry_vm_cl1,entry_cl2,entry_vm_cl2,entry_cl3,entry_vm_cl3,
    entry_cl4,entry_vm_cl4,entry_cl5,entry_vm_cl5,entry_cl6,entry_vm_cl6,
    entry_tp1,entry_vm_tp1,entry_tp2,entry_vm_tp2,entry_tp3,entry_vm_tp3,
    entry_tp4,entry_vm_tp4,entry_tp5,entry_vm_tp5,entry_tp6,entry_vm_tp6,
    entry_path)
    root.geometry('800x600')
    root.mainloop()

def change(entry_ca_con,entry_alk_con,entry_cl_con,entry_tp_con):
    with open('standard_concentration.txt','w') as f:
        ca_hn_value = float(entry_ca_con.get())
        alk_value = float(entry_alk_con.get())
        cl_value = float(entry_cl_con.get())
        tp_value = float(entry_tp_con.get())
        s = str(ca_hn_value)+','+str(alk_value)+','+str(cl_value)+','+str(tp_value)
        f.write(s)

def cclt(entry_water1,entry_water2,entry_water3,entry_water4,entry_water5,entry_water6,
    entry_ca1,entry_vm_ca1,entry_ca2,entry_vm_ca2,entry_ca3,entry_vm_ca3,
    entry_ca4,entry_vm_ca4,entry_ca5,entry_vm_ca5,entry_ca6,entry_vm_ca6,
    entry_hn1,entry_vm_hn1,entry_hn2,entry_vm_hn2,entry_hn3,entry_vm_hn3,
    entry_hn4,entry_vm_hn4,entry_hn5,entry_vm_hn5,entry_hn6,entry_vm_hn6,
    entry_alk1,entry_vm_alk1,entry_alk2,entry_vm_alk2,entry_alk3,entry_vm_alk3,
    entry_alk4,entry_vm_alk4,entry_alk5,entry_vm_alk5,entry_alk6,entry_vm_alk6,
    entry_cl1,entry_vm_cl1,entry_cl2,entry_vm_cl2,entry_cl3,entry_vm_cl3,
    entry_cl4,entry_vm_cl4,entry_cl5,entry_vm_cl5,entry_cl6,entry_vm_cl6,
    entry_tp1,entry_vm_tp1,entry_tp2,entry_vm_tp2,entry_tp3,entry_vm_tp3,
    entry_tp4,entry_vm_tp4,entry_tp5,entry_vm_tp5,entry_tp6,entry_vm_tp6,
    entry_path):
    
    with open('standard_concentration.txt','r') as f:
        for line in f.readlines():
            l = line.split(',')
            float(l[0])
            ca_hn_value = float(l[0])
            alk_value = float(l[1])
            cl_value = float(l[2])
            tp_value = float(l[3])
    
    # Get key water items
    key_water1 = str(entry_water1.get())
    key_water2 = str(entry_water2.get())
    key_water3 = str(entry_water3.get())
    key_water4 = str(entry_water4.get())
    key_water5 = str(entry_water5.get())
    key_water6 = str(entry_water6.get())

    # Get ca value 
    value_ca1 = round(float(entry_ca1.get())*ca_hn_value*100*1000/float(entry_vm_ca1.get()))
    value_ca2 = round(float(entry_ca2.get())*ca_hn_value*100*1000/float(entry_vm_ca2.get()))
    value_ca3 = round(float(entry_ca3.get())*ca_hn_value*100*1000/float(entry_vm_ca3.get()))
    value_ca4 = round(float(entry_ca4.get())*ca_hn_value*100*1000/float(entry_vm_ca4.get()))
    value_ca5 = round(float(entry_ca5.get())*ca_hn_value*100*1000/float(entry_vm_ca5.get()))
    value_ca6 = round(float(entry_ca6.get())*ca_hn_value*100*1000/float(entry_vm_ca6.get()))

    # Get hardness value 
    value_hn1 = round(float(entry_hn1.get())*ca_hn_value*100*1000/float(entry_vm_hn1.get()))
    value_hn2 = round(float(entry_hn2.get())*ca_hn_value*100*1000/float(entry_vm_hn2.get()))
    value_hn3 = round(float(entry_hn3.get())*ca_hn_value*100*1000/float(entry_vm_hn3.get()))
    value_hn4 = round(float(entry_hn4.get())*ca_hn_value*100*1000/float(entry_vm_hn4.get()))
    value_hn5 = round(float(entry_hn5.get())*ca_hn_value*100*1000/float(entry_vm_hn5.get()))
    value_hn6 = round(float(entry_hn6.get())*ca_hn_value*100*1000/float(entry_vm_hn6.get()))

    # Get alk value 
    value_alk1 = round(float(entry_alk1.get())*alk_value*50*1000/float(entry_vm_alk1.get()))
    value_alk2 = round(float(entry_alk2.get())*alk_value*50*1000/float(entry_vm_alk2.get()))
    value_alk3 = round(float(entry_alk3.get())*alk_value*50*1000/float(entry_vm_alk3.get()))
    value_alk4 = round(float(entry_alk4.get())*alk_value*50*1000/float(entry_vm_alk4.get()))
    value_alk5 = round(float(entry_alk5.get())*alk_value*50*1000/float(entry_vm_alk5.get()))
    value_alk6 = round(float(entry_alk6.get())*alk_value*50*1000/float(entry_vm_alk6.get()))

    # Get cl value 
    value_cl1 = round(float(entry_cl1.get())*cl_value*0.35453*100*1000/float(entry_vm_cl1.get()))
    value_cl2 = round(float(entry_cl2.get())*cl_value*0.35453*100*1000/float(entry_vm_cl2.get()))
    value_cl3 = round(float(entry_cl3.get())*cl_value*0.35453*100*1000/float(entry_vm_cl3.get()))
    value_cl4 = round(float(entry_cl4.get())*cl_value*0.35453*100*1000/float(entry_vm_cl4.get()))
    value_cl5 = round(float(entry_cl5.get())*cl_value*0.35453*100*1000/float(entry_vm_cl5.get()))
    value_cl6 = round(float(entry_cl6.get())*cl_value*0.35453*100*1000/float(entry_vm_cl6.get()))

    # Get total phosphorus 
    value_tp1 = round(float(entry_tp1.get())*1000*tp_value/float(entry_vm_tp1.get()),2)
    value_tp2 = round(float(entry_tp2.get())*1000*tp_value/float(entry_vm_tp2.get()),2)
    value_tp3 = round(float(entry_tp3.get())*1000*tp_value/float(entry_vm_tp3.get()),2)
    value_tp4 = round(float(entry_tp4.get())*1000*tp_value/float(entry_vm_tp4.get()),2)
    value_tp5 = round(float(entry_tp5.get())*1000*tp_value/float(entry_vm_tp5.get()),2)
    value_tp6 = round(float(entry_tp6.get())*1000*tp_value/float(entry_vm_tp6.get()),2)
    
    names = [key_water1,key_water2,key_water3,key_water4,key_water5,key_water6]
    data_1 = [value_ca1,value_hn1,value_alk1,value_cl1,value_tp1,0]
    data_2 = [value_ca2,value_hn2,value_alk2,value_cl2,value_tp2,1]
    data_3 = [value_ca3,value_hn3,value_alk3,value_cl3,value_tp3,2]
    data_4 = [value_ca4,value_hn4,value_alk4,value_cl4,value_tp4,3]
    data_5 = [value_ca5,value_hn5,value_alk5,value_cl5,value_tp5,4]
    data_6 = [value_ca6,value_hn6,value_alk6,value_cl6,value_tp6,5]
    
    wb = load_workbook('Report_Template.xlsx')
    ws = wb.active
    ws.title = str(entry_path.get()) + 'Water Analysis Report'
    today = datetime.date.today()
    today = str(today)
    today_date = ws.cell(column=6,row=4,value=today)
    today_date.alignment = Alignment(horizontal='center',vertical='center')
    num_names = 0
    for i in names:
        if len(i)>1:
            num_names +=1
    num_water = ws.cell(column=6,row=3,value=num_names)
    num_water.alignment = Alignment(horizontal='center',vertical='bottom')
    def write_names(names):
        i = 0
        for column in range(3,3+len(names)):
            if names[i] != '0':
                _ = ws.cell(column=column,row = 6,value=names[i])
                _.alignment = Alignment(horizontal='center',vertical='center')
            i += 1
    def write_values(data):
        i = 0
        col = 3 + int(list(data)[-1])
        for row in range(12,22,2):
            if data[i] != 0:
                _ = ws.cell(column = col,row=row,value=data[i])
                _.alignment = Alignment(horizontal='center',vertical='bottom')
            i += 1
    write_names(names)
    write_values(data_1)
    write_values(data_2)
    write_values(data_3) 
    write_values(data_4)    
    write_values(data_5) 
    write_values(data_6)
    data = [data_1,data_2,data_3,data_4,data_5,data_6]
    ws['C2'] = str(entry_path.get())
    project_name = str(entry_path.get())
    new_xlsx_name = today+str(entry_path.get())+'water_analysis_report'+'.xlsx'
    wb.save(new_xlsx_name)
    
    database = 'WaterSamplesDB.db'
    connection = sqlite3.connect(database,timeout= 10)
    cursor = connection.cursor()
    #CREATE WaterType TABLE
    cursor.execute('''CREATE TABLE IF NOT EXISTS WaterType(
        TypeId integer PRIMARY KEY,
        TypeName Text NOT NULL)''')

    #CREATE TestItems TABLE
    cursor.execute('''CREATE TABLE IF NOT EXISTS TestItems(
        TestId integer PRIMARY KEY,
        TestName text NOT NULL)''')

    #CREATE Reports TABLE
    cursor.execute('''CREATE TABLE IF NOT EXISTS Reports(
        ProjectName text NOT NULL,
        Date text NOT NULL,
        TestId integer,
        TypeId integer,
        Data numeric) ''')
    connection.commit()

    #INSERT INTO An Example Sample
    Type_list = [
        (1,'Circulation Water'),
        (2,'Make Up'),
        (3,'Return Water'),
        (4,'Sewage'),
        (5,'RO')]

    TestItems_list = [
        (1,'Ca'),
        (2,'Harness'),
        (3,'Alk'),
        (4,'Cl'),
        (5,'TP')]

    cursor.execute("select * from WaterType where TypeId = 1")
    if len(cursor.fetchall()) == 0:
        cursor.executemany("INSERT INTO WaterType VALUES (?,?)",Type_list)
        connection.commit()

    cursor.execute("select * from TestItems where TestId = 1")
    if len(cursor.fetchall()) == 0:
        cursor.executemany("INSERT INTO TestItems VALUES(?,?)",TestItems_list)
        connection.commit()
    
    connection.close()
    
    Type_id = 0

    for n in range(len(names)):
        if names[n] != '0':
            if names[n] == 'Circulation Water':
                Type_id = 1
            elif names[n] == 'Make up':
                Type_id = 2
            elif names[n] == 'Return Water':
                Type_id = 3
            elif names[n] == 'Sewage':
                Type_id = 4
            elif names[n] == 'RO':
                Type_id = 5
            for d in range(len(data[n])-1):
                if data[n][d] != 0:
                    database = 'WaterSamplesDB.db'
                    connection = sqlite3.connect(database,timeout= 1)
                    cursor = connection.cursor()
                    sql = ("INSERT OR REPLACE INTO Reports (ProjectName, Date, TestId, TypeId, Data) VALUES (?,?,?,?,?)")
                    val = (project_name, today, d, Type_id, data[n][d])
                    cursor.execute(sql, val)
                    connection.commit()
                    connection.close()


    print('project_name: ',project_name)
    print('Date: ',today)
    print('Data: ',data)
    print('Name: ',names)

    
main()