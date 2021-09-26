import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
names = ['水样1','水样2','水样3']
data_1 = [1,2,3,4,5,0]
data_2 = [2,3,4,5,6,1]
data_3 = [3,4,5,6,7,2]

wb2 = load_workbook('test_ex.xlsx')
ws1 = wb2.active
def write_names(names):
    i = 0
    for column in range(3,3+len(names)):
        _ = ws1.cell(column=column,row = 3,value=names[i])
        i += 1
def write_values(data):
    i = 0
    col = 3 + int(list(data)[-1])
    for row in range(12,22,2):
        _ = ws1.cell(column = col,row=row,value=data[i])
        i += 1

write_names(names)
write_values(data_1)
#write_values(data_2)
#write_values(data_3)
#i_1 = 0
#for row in range(12,22,2):
#   _ = ws1.cell(column = 3,row=row,value=l1[i_1])
#    i_1 += 1
wb2.save('test_ex.xlsx')

alignment = Alignment(vertical=...)